import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
import plotly.express as px
from streamlit_option_menu import option_menu
from PIL import Image
from datetime import datetime

# ------------------------------
# CONFIGURATION
# ------------------------------
st.set_page_config(page_title="Berber Cement KPI Dashboard", layout="wide", page_icon="🏭")

# Note: Update this path for deployment or make it configurable
DATA_FOLDER = Path(r"C:\Users\amar.sirajeldin\Desktop\Plant parameters")
MAIN_KPI_FILE = DATA_FOLDER / "Berber_Cement_KPIs_Dashboard.xlsx"
VRM_ADVANCED_FILE = DATA_FOLDER / "berber_vrm_ml_worksheet_advanced.xlsx"
KILN_ML_FILE = DATA_FOLDER / "kiln_ml_worksheet_example.xlsx"
VRM_ML_FILE = DATA_FOLDER / "vrm_ml_worksheet_example.xlsx"
LOGO_PATH = DATA_FOLDER / "Logo.jpg"

# ------------------------------
# UTILITY FUNCTIONS
# ------------------------------
@st.cache_data
def load_workbook_sheets(file_path):
    """Load all sheets from an Excel file"""
    if not file_path.exists():
        st.error(f"File not found: {file_path}")
        return {}
    xl = pd.ExcelFile(file_path)
    sheets = {}
    for sheet in xl.sheet_names:
        df = xl.parse(sheet, header=None)
        sheets[sheet] = df
    return sheets

@st.cache_data
def load_model_config(file_path, sheet_name="Model_Config"):
    """Load model coefficients from Model_Config sheet"""
    sheets = load_workbook_sheets(file_path)
    if sheet_name not in sheets:
        return {}
    
    df = sheets[sheet_name]
    config = {}
    
    for i, row in df.iterrows():
        if len(row) >= 2 and pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
            param = str(row.iloc[0]).strip()
            value = row.iloc[1]
            # Try to convert to numeric if possible
            try:
                if isinstance(value, str):
                    # Check if it's a number
                    value = float(value) if value.replace('.', '').replace('-', '').isdigit() else value
                config[param] = value
            except:
                config[param] = value
    return config

def parse_kpi_sheet(df):
    """Improved parser that detects KPI table start reliably."""
    start_row = None
    for i, row in df.iterrows():
        val_a = row.iloc[0] if len(row) > 0 else None
        val_b = row.iloc[1] if len(row) > 1 else None
        if val_a and isinstance(val_a, str):
            if "Back to Home" in val_a or "📊" in val_a or "🏠" in val_a:
                continue
            try:
                pd.to_numeric(val_b)
                start_row = i
                break
            except:
                pass
    if start_row is None:
        return pd.DataFrame()

    end_row = min(start_row + 30, len(df))
    table = df.iloc[start_row:end_row].copy()
    table.columns = ["KPI", "Target", "Actual", "Unit", "E", "F", "G"][:len(table.columns)]
    table = table[["KPI", "Target", "Actual", "Unit"]]

    table = table.dropna(subset=["KPI"])
    table = table[~table["KPI"].astype(str).str.contains("Back to Home|📊|🏠|Remarks", na=False)]
    table = table[table["KPI"].astype(str).str.strip() != ""]

    table["Target"] = pd.to_numeric(table["Target"], errors="coerce")
    table["Actual"] = pd.to_numeric(table["Actual"], errors="coerce")
    table = table.dropna(subset=["Target", "Actual"], how="all")
    return table.reset_index(drop=True)

def compute_status(kpi_name, target, actual):
    if pd.isna(actual) or pd.isna(target):
        return ""
    lower_better_kw = [
        "consumption", "cost", "downtime", "vibration", "leakage", "emissions",
        "index", "time", "temperature", "loading", "content", "free lime",
        "ratio", "distance", "moisture", "size", "cases", "noise", "response time",
        "mttr", "scan time", "response time", "leakage", "wear life", "deviation",
        "risk", "specific"
    ]
    is_lower_better = any(kw in kpi_name.lower() for kw in lower_better_kw)

    if is_lower_better:
        if actual <= target:
            return "✅ On Track"
        elif actual <= target * 1.05:
            return "⚠️ Attention"
        else:
            return "❌ Critical"
    else:
        if actual >= target:
            return "✅ On Track"
        elif actual >= target * 0.95:
            return "⚠️ Attention"
        else:
            return "❌ Critical"

def save_actuals_to_excel(file_path, sheet_name, edited_df):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    start_row = None
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        cell_b = ws.cell(row=row, column=2).value
        if cell_a and isinstance(cell_a, str):
            if "Back to Home" in cell_a or "📊" in cell_a:
                continue
            try:
                float(cell_b)
                start_row = row
                break
            except:
                pass
    if start_row is None:
        return

    for i, actual in enumerate(edited_df["Actual"]):
        ws.cell(row=start_row + i, column=3, value=actual)

    wb.save(file_path)

# ------------------------------
# KILN ML FUNCTIONS
# ------------------------------
def calculate_kiln_predictions(live_values, config):
    """Calculate kiln predictions based on model coefficients"""
    # Calculate deviations
    o2_dev = abs(live_values.get('o2_percent', 0) - config.get('Target back-end O2 %', 2.2))
    calciner_dev = abs(live_values.get('calciner_temp_c', 0) - config.get('Target calciner temp C', 880)) / 10
    inlet_dev = abs(live_values.get('kiln_inlet_temp_c', 0) - config.get('Target kiln inlet temp C', 1050)) / 10
    coal_feed_ratio = live_values.get('coal_tph', 0) / live_values.get('feed_tph', 1) if live_values.get('feed_tph', 0) > 0 else 0
    ratio_dev = abs(coal_feed_ratio - config.get('Target coal/feed ratio', 0.108)) * 100
    draft_dev = abs(live_values.get('draft_pa', 0) - config.get('Target draft Pa', -5200)) / 100
    feed_delta = 0  # For live scoring, feed delta is 0 (single point)
    sat_deficit = max(0, config.get('Target secondary air temp C', 1040) - live_values.get('sat_c', 0)) / 10
    cooler_high = max(0, live_values.get('cooler_out_c', 0) - config.get('Target cooler outlet temp C', 165)) / 10
    
    # Pred Free Lime (linear model)
    pred_free_lime = (
        config.get('FL intercept', 0.25) +
        config.get('FL O2 dev', 0.18) * o2_dev +
        config.get('FL calciner dev/10', 0.08) * calciner_dev +
        config.get('FL inlet dev/10', 0.06) * inlet_dev +
        config.get('FL ratio dev x100', 0.05) * ratio_dev +
        config.get('FL draft dev/100', 0.03) * draft_dev +
        config.get('FL feed delta/10', 0.04) * feed_delta +
        config.get('FL SAT deficit/10', 0.06) * sat_deficit
    )
    
    # Ring Risk Probability (logistic model)
    logit = (
        config.get('Ring intercept', -4) +
        config.get('Ring O2 dev', 0.9) * o2_dev +
        config.get('Ring draft dev/100', 0.55) * draft_dev +
        config.get('Ring SAT deficit/10', 0.45) * sat_deficit +
        config.get('Ring cooler high/10', 0.35) * cooler_high +
        config.get('Ring feed delta/10', 0.5) * feed_delta
    )
    ring_risk = 1 / (1 + np.exp(-logit))
    
    # Stability Score
    stability_score = max(0, 100 - (
        config.get('Stab O2 weight', 6) * o2_dev +
        config.get('Stab calciner weight', 4) * calciner_dev +
        config.get('Stab inlet weight', 4) * inlet_dev +
        config.get('Stab draft weight', 3) * draft_dev +
        config.get('Stab feed delta weight', 5) * feed_delta +
        config.get('Stab SAT deficit weight', 4) * sat_deficit
    ))
    
    return {
        'o2_deviation': o2_dev,
        'calciner_dev_10': calciner_dev,
        'inlet_dev_10': inlet_dev,
        'coal_feed_ratio': coal_feed_ratio,
        'ratio_dev_x100': ratio_dev,
        'draft_dev_100': draft_dev,
        'sat_deficit_10': sat_deficit,
        'cooler_high_10': cooler_high,
        'pred_free_lime': pred_free_lime,
        'ring_risk_probability': ring_risk,
        'stability_score': stability_score
    }

def render_kiln_ml():
    st.header("🔥 Kiln ML Scorer - Advanced Analytics")
    
    if not KILN_ML_FILE.exists():
        st.error(f"Kiln ML file not found at: {KILN_ML_FILE}")
        return
    
    # Load model configuration
    config = load_model_config(KILN_ML_FILE, "Model_Config")
    if not config:
        st.error("Could not load model configuration from Kiln ML file")
        return
    
    st.markdown("### Live Kiln Parameters")
    st.markdown("*Enter current kiln operating parameters for real-time predictions*")
    
    col1, col2 = st.columns(2)
    
    with col1:
        feed_tph = st.number_input("Kiln Feed (tph)", value=317.5, step=5.0, format="%.1f", key="kiln_feed")
        calciner_temp = st.number_input("Calciner Temperature (°C)", value=873.4, step=5.0, format="%.1f", key="kiln_calciner")
        inlet_temp = st.number_input("Kiln Inlet Temperature (°C)", value=1041.5, step=5.0, format="%.1f", key="kiln_inlet")
        o2_percent = st.number_input("Back-end O₂ (%)", value=1.99, step=0.1, format="%.2f", key="kiln_o2")
    
    with col2:
        sat_temp = st.number_input("Secondary Air Temperature (°C)", value=1049.8, step=5.0, format="%.1f", key="kiln_sat")
        cooler_out = st.number_input("Cooler Outlet Temperature (°C)", value=156.3, step=5.0, format="%.1f", key="kiln_cooler")
        coal_tph = st.number_input("Coal Feed (tph)", value=34.3, step=1.0, format="%.1f", key="kiln_coal")
        draft_pa = st.number_input("Draft Pressure (Pa)", value=-5291, step=50, format="%.0f", key="kiln_draft")
    
    live_values = {
        'feed_tph': feed_tph,
        'calciner_temp_c': calciner_temp,
        'kiln_inlet_temp_c': inlet_temp,
        'o2_percent': o2_percent,
        'sat_c': sat_temp,
        'cooler_out_c': cooler_out,
        'coal_tph': coal_tph,
        'draft_pa': draft_pa
    }
    
    # Calculate predictions
    predictions = calculate_kiln_predictions(live_values, config)
    
    st.markdown("---")
    st.markdown("### ML Model Predictions")
    
    # Display predictions in metrics
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            "Predicted Free Lime (%)",
            f"{predictions['pred_free_lime']:.2f}",
            delta=f"Target: {config.get('Target free lime %', 1.2)}%",
            delta_color="inverse"
        )
        st.metric(
            "Ring Risk Probability",
            f"{predictions['ring_risk_probability']:.1%}",
            delta=f"Threshold: {config.get('Ring risk threshold', 0.7):.0%}",
            delta_color="inverse"
        )
    
    with col2:
        st.metric(
            "Stability Score",
            f"{predictions['stability_score']:.0f}",
            delta=f"Floor: {config.get('Stability score floor', 80)}",
            delta_color="inverse"
        )
        
        # Determine overall status
        if predictions['ring_risk_probability'] >= config.get('Ring risk threshold', 0.7):
            status = "🔴 HIGH RING RISK"
            status_color = "red"
        elif predictions['pred_free_lime'] >= config.get('Pred free lime threshold', 1.8):
            status = "🟠 QUALITY RISK"
            status_color = "orange"
        elif predictions['stability_score'] < config.get('Stability score floor', 80):
            status = "🟡 CAUTION"
            status_color = "yellow"
        else:
            status = "🟢 NORMAL"
            status_color = "green"
        
        st.markdown(f"**Overall Status:** <span style='color:{status_color}'>{status}</span>", unsafe_allow_html=True)
    
    with col3:
        st.metric("O₂ Deviation", f"{predictions['o2_deviation']:.2f}%")
        st.metric("Coal/Feed Ratio", f"{predictions['coal_feed_ratio']:.3f}")
    
    # Detailed metrics expander
    with st.expander("📊 Detailed Deviation Metrics"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Calciner Deviation/10", f"{predictions['calciner_dev_10']:.2f}")
            st.metric("Inlet Deviation/10", f"{predictions['inlet_dev_10']:.2f}")
        with col2:
            st.metric("Ratio Deviation x100", f"{predictions['ratio_dev_x100']:.2f}")
            st.metric("Draft Deviation/100", f"{predictions['draft_dev_100']:.2f}")
        with col3:
            st.metric("SAT Deficit/10", f"{predictions['sat_deficit_10']:.2f}")
            st.metric("Cooler High/10", f"{predictions['cooler_high_10']:.2f}")
    
    # Suggested actions
    st.markdown("---")
    st.markdown("### 💡 Suggested Actions")
    
    if status == "🔴 HIGH RING RISK":
        st.warning("⚠️ **HIGH RING RISK DETECTED**")
        st.markdown("""
        - Check draft stability and coating/ring tendency
        - Monitor cooler performance and SAT temperature
        - Review O₂ control and combustion conditions
        - Consider reducing feed rate temporarily
        - Schedule inspection for potential ring formation
        """)
    elif status == "🟠 QUALITY RISK":
        st.info("📊 **QUALITY RISK DETECTED**")
        st.markdown("""
        - Review burning intensity and flame shape
        - Adjust O₂ control and coal/feed ratio
        - Check calciner temperature stability
        - Monitor free lime trend closely
        - Consider raw mix adjustment if persistent
        """)
    elif status == "🟡 CAUTION":
        st.info("⚠️ **OPERATION CAUTION**")
        st.markdown("""
        - Monitor stability parameters closely
        - Check for feed rate fluctuations
        - Review draft and temperature trends
        - Prepare for potential intervention if degradation continues
        """)
    else:
        st.success("✅ **NORMAL OPERATION**")
        st.markdown("Continue monitoring and maintain current operating parameters.")

# ------------------------------
# VRM ML FUNCTIONS
# ------------------------------
def calculate_vrm_predictions(live_values, config):
    """Calculate VRM predictions based on model coefficients"""
    # Calculate deviations
    vibration_dev = abs(live_values.get('vibration', 0) - config.get('Target vibration mm/s', 1.6))
    dp_dev = abs(live_values.get('mill_dp', 0) - config.get('Target mill DP mbar', 78)) / 10
    temp_dev = abs(live_values.get('outlet_temp', 0) - config.get('Target outlet temp C', 86)) / 10
    separator_dev = abs(live_values.get('separator_rpm', 0) - config.get('Target separator rpm', 930)) / 100
    reject_dev = abs(live_values.get('reject_tph', 0) - config.get('Target reject tph', 14))
    fan_dev = abs(live_values.get('fan_damper', 0) - config.get('Target fan damper %', 72)) / 10
    feed_delta = 0  # For live scoring, feed delta is 0
    
    # Predicted Residue (linear model)
    pred_residue = (
        config.get('Residue intercept', 5.2) +
        config.get('Coeff vibration dev', 0.95) * vibration_dev +
        config.get('Coeff DP dev/10', 0.42) * dp_dev +
        config.get('Coeff outlet temp dev/10', 0.28) * temp_dev +
        config.get('Coeff reject dev', 0.18) * reject_dev +
        config.get('Coeff fan dev/10', 0.11) * fan_dev +
        config.get('Coeff feed delta/10', 0.16) * feed_delta +
        config.get('Coeff separator dev/100', 0.24) * separator_dev
    )
    
    # Trip Risk Probability (logistic model)
    logit = (
        config.get('Trip intercept', -4.1) +
        config.get('Coeff trip vibration dev', 1.55) * vibration_dev +
        config.get('Coeff trip reject dev', 0.12) * reject_dev +
        config.get('Coeff trip feed delta/10', 0.48) * feed_delta +
        config.get('Coeff trip DP dev/10', 0.65) * dp_dev +
        config.get('Coeff trip fan dev/10', 0.2) * fan_dev
    )
    trip_risk = 1 / (1 + np.exp(-logit))
    
    # Stability Score
    stability_score = max(0, 100 - (
        config.get('Stability wt vibration dev', 18) * vibration_dev +
        config.get('Stability wt DP dev/10', 8) * dp_dev +
        config.get('Stability wt outlet temp dev/10', 6) * temp_dev +
        config.get('Stability wt reject dev', 1.2) * reject_dev +
        config.get('Stability wt fan dev/10', 3.5) * fan_dev +
        config.get('Stability wt feed delta/10', 4) * feed_delta
    ))
    
    return {
        'vibration_deviation': vibration_dev,
        'dp_dev_10': dp_dev,
        'temp_dev_10': temp_dev,
        'separator_dev_100': separator_dev,
        'reject_deviation': reject_dev,
        'fan_dev_10': fan_dev,
        'pred_residue': pred_residue,
        'trip_risk_probability': trip_risk,
        'stability_score': stability_score
    }

def render_vrm_advanced():
    st.header("📈 VRM Advanced ML Scorer")
    
    if not VRM_ML_FILE.exists():
        st.error(f"VRM ML file not found at: {VRM_ML_FILE}")
        return
    
    # Load model configuration
    config = load_model_config(VRM_ML_FILE, "Model_Config")
    if not config:
        st.error("Could not load model configuration from VRM ML file")
        return
    
    st.markdown("### Live VRM Parameters")
    st.markdown("*Enter current VRM operating parameters for real-time predictions*")
    
    col1, col2 = st.columns(2)
    
    with col1:
        feed_tph = st.number_input("Feed (tph)", value=183.8, step=5.0, format="%.1f", key="vrm_feed")
        mill_dp = st.number_input("Mill DP (mbar)", value=79.4, step=2.0, format="%.1f", key="vrm_dp")
        outlet_temp = st.number_input("Outlet Temperature (°C)", value=84.9, step=1.0, format="%.1f", key="vrm_temp")
        vibration = st.number_input("Vibration (mm/s)", value=1.72, step=0.1, format="%.2f", key="vrm_vibration")
    
    with col2:
        separator_rpm = st.number_input("Separator Speed (rpm)", value=925, step=10, format="%.0f", key="vrm_separator")
        reject_tph = st.number_input("Reject (tph)", value=15.2, step=1.0, format="%.1f", key="vrm_reject")
        main_motor_kw = st.number_input("Main Motor Power (kW)", value=5090, step=50, format="%.0f", key="vrm_motor")
        fan_damper = st.number_input("Fan Damper (%)", value=73.6, step=1.0, format="%.1f", key="vrm_fan")
    
    live_values = {
        'feed_tph': feed_tph,
        'mill_dp': mill_dp,
        'outlet_temp': outlet_temp,
        'vibration': vibration,
        'separator_rpm': separator_rpm,
        'reject_tph': reject_tph,
        'main_motor_kw': main_motor_kw,
        'fan_damper': fan_damper
    }
    
    # Calculate predictions
    predictions = calculate_vrm_predictions(live_values, config)
    
    st.markdown("---")
    st.markdown("### ML Model Predictions")
    
    # Display predictions in metrics
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            "Predicted Residue (%)",
            f"{predictions['pred_residue']:.1f}",
            delta=f"Limit: {config.get('Residue prediction limit %', 12.5)}%",
            delta_color="inverse"
        )
        st.metric(
            "Trip Risk Probability",
            f"{predictions['trip_risk_probability']:.1%}",
            delta=f"Threshold: {config.get('Trip risk probability threshold', 0.65):.0%}",
            delta_color="inverse"
        )
    
    with col2:
        st.metric(
            "Stability Score",
            f"{predictions['stability_score']:.0f}",
            delta=f"Minimum: {config.get('Minimum stability score', 70)}",
            delta_color="inverse"
        )
        
        # Determine overall status
        if predictions['trip_risk_probability'] >= config.get('Trip risk probability threshold', 0.65):
            status = "🔴 HIGH TRIP RISK"
            status_color = "red"
        elif predictions['pred_residue'] >= config.get('Residue prediction limit %', 12.5):
            status = "🟠 QUALITY RISK"
            status_color = "orange"
        elif predictions['stability_score'] < config.get('Minimum stability score', 70):
            status = "🟡 CAUTION"
            status_color = "yellow"
        else:
            status = "🟢 NORMAL"
            status_color = "green"
        
        st.markdown(f"**Overall Status:** <span style='color:{status_color}'>{status}</span>", unsafe_allow_html=True)
    
    with col3:
        st.metric("Vibration Deviation", f"{predictions['vibration_deviation']:.2f} mm/s")
        st.metric("Reject Deviation", f"{predictions['reject_deviation']:.1f} tph")
    
    # Detailed metrics expander
    with st.expander("📊 Detailed Deviation Metrics"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("DP Deviation/10", f"{predictions['dp_dev_10']:.2f}")
            st.metric("Temperature Deviation/10", f"{predictions['temp_dev_10']:.2f}")
        with col2:
            st.metric("Separator Deviation/100", f"{predictions['separator_dev_100']:.2f}")
            st.metric("Fan Damper Deviation/10", f"{predictions['fan_dev_10']:.2f}")
        with col3:
            specific_power = main_motor_kw / feed_tph if feed_tph > 0 else 0
            st.metric("Specific Power", f"{specific_power:.1f} kWh/t")
            st.metric("Target Specific Power", f"{config.get('Target specific power kWh/t', 27.5):.1f} kWh/t")
    
    # Suggested actions
    st.markdown("---")
    st.markdown("### 💡 Suggested Actions")
    
    if status == "🔴 HIGH TRIP RISK":
        st.warning("⚠️ **HIGH TRIP RISK DETECTED**")
        st.markdown("""
        - Reduce feed rate fluctuations immediately
        - Inspect vibration trend and source
        - Check reject rate and circulating load
        - Monitor mill DP for bed instability
        - Consider reducing feed rate temporarily
        """)
    elif status == "🟠 QUALITY RISK":
        st.info("📊 **QUALITY RISK DETECTED**")
        st.markdown("""
        - Review separator speed and classification efficiency
        - Adjust grinding pressure and dam ring height
        - Check circulating load and reject rate
        - Monitor product fineness trend
        - Consider mill ventilation adjustment
        """)
    elif status == "🟡 CAUTION":
        st.info("⚠️ **OPERATION CAUTION**")
        st.markdown("""
        - Monitor stability parameters closely
        - Watch for increasing vibration trend
        - Check feed consistency
        - Prepare for potential intervention if degradation continues
        """)
    else:
        st.success("✅ **NORMAL OPERATION**")
        st.markdown("Continue monitoring and maintain current operating parameters.")

# ------------------------------
# PAGE RENDERERS
# ------------------------------
def render_kpi_page(sheet_name, title, icon=""):
    st.header(f"{icon} {title}")
    sheets = load_workbook_sheets(MAIN_KPI_FILE)
    if sheet_name not in sheets:
        st.error(f"Sheet '{sheet_name}' not found.")
        return

    df_raw = sheets[sheet_name]
    kpi_df = parse_kpi_sheet(df_raw)
    if kpi_df.empty:
        st.warning("No KPIs found in this sheet.")
        return

    kpi_df["Status"] = kpi_df.apply(
        lambda r: compute_status(r["KPI"], r["Target"], r["Actual"]), axis=1
    )

    edited_df = st.data_editor(
        kpi_df,
        column_config={
            "KPI": st.column_config.TextColumn(disabled=True),
            "Target": st.column_config.NumberColumn(disabled=True, format="%.2f"),
            "Actual": st.column_config.NumberColumn(format="%.2f"),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Status": st.column_config.TextColumn(disabled=True),
        },
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        key=f"editor_{sheet_name}"
    )

    edited_df["Status"] = edited_df.apply(
        lambda r: compute_status(r["KPI"], r["Target"], r["Actual"]), axis=1
    )

    col1, col2, col3 = st.columns(3)
    on_track = (edited_df["Status"] == "✅ On Track").sum()
    attention = (edited_df["Status"] == "⚠️ Attention").sum()
    critical = (edited_df["Status"] == "❌ Critical").sum()
    col1.metric("✅ On Track", on_track)
    col2.metric("⚠️ Attention", attention)
    col3.metric("❌ Critical", critical)

    chart_data = edited_df.dropna(subset=["Target", "Actual"]).copy()
    if not chart_data.empty:
        fig = px.bar(chart_data.melt(id_vars=["KPI"], value_vars=["Target", "Actual"]),
                     x="KPI", y="value", color="variable", barmode="group",
                     title="Target vs Actual", height=400)
        st.plotly_chart(fig, use_container_width=True)

    if st.button("💾 Save Changes to Excel", key=f"save_{sheet_name}"):
        save_actuals_to_excel(MAIN_KPI_FILE, sheet_name, edited_df)
        st.success("Saved successfully! Refresh to see updated data.")
        st.cache_data.clear()

def render_maintenance_subpages():
    subpage = st.selectbox("Select Maintenance Section", 
                           ["Mechanical", "Electrical", "DCS & Instrument", "Heavy Equipment"])
    sheet_map = {
        "Mechanical": "MAINTENANCE - MECHANICAL",
        "Electrical": "MAINTENANCE - ELECTRICAL",
        "DCS & Instrument": "MAINTENANCE - DCS & INSTRUMENT",
        "Heavy Equipment": "MAINTENANCE - HEAVY EQUIPMENT"
    }
    render_kpi_page(sheet_map[subpage], f"Maintenance - {subpage}", "🔧")

def render_utility_subpages():
    subpage = st.selectbox("Select Utility Section", 
                           ["Civil", "Industrial Services", "HSE"])
    sheet_map = {
        "Civil": "UTILITY - CIVIL",
        "Industrial Services": "UTILITY - INDUSTRIAL SERVICES",
        "HSE": "HSE"
    }
    render_kpi_page(sheet_map[subpage], f"Utility - {subpage}", "🔨")

# ------------------------------
# MAIN APP
# ------------------------------
def main():
    # Sidebar with small centered logo
    with st.sidebar:
        if LOGO_PATH.exists():
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                logo = Image.open(LOGO_PATH)
                st.image(logo, width=150)
        else:
            st.warning("Logo not found at: " + str(LOGO_PATH))
        st.title("Berber Cement")
        selected = option_menu(
            menu_title="Navigation",
            options=[
                "Home",
                "Production",
                "Maintenance",
                "Utility",
                "Power Generation",
                "Quality Control",
                "Quarry & Crusher",
                "VRM Advanced ML",
                "Kiln ML"
            ],
            icons=["house", "speedometer2", "wrench", "tools", "lightning", "clipboard-check", "minecart", "graph-up", "fire"],
            menu_icon="cast",
            default_index=0,
        )

    # Page routing
    if selected == "Home":
        st.title("🏭 Berber Cement Plant KPI Dashboard")
        st.markdown("### Welcome to the interactive KPI platform")
        st.markdown("""
        - **Navigate** using the sidebar menu.
        - **Edit** actual values directly in the tables.
        - **Save** changes to Excel with the button at the bottom.
        - **Status colors**: ✅ On Track, ⚠️ Attention, ❌ Critical.
        - **ML Models**: Use VRM and Kiln ML scorers for predictive analytics.
        """)
        st.info(f"Data folder: `{DATA_FOLDER}`")
        st.caption(f"Last updated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Production", "6,000 t/d", "Target")
        col2.metric("Kiln Feed", "400 tph", "Target")
        col3.metric("Sp. Heat", "720 kcal/kg", "Target")
        col4.metric("OEE", "85%", "Target")

    elif selected == "Production":
        render_kpi_page("PRODUCTION", "Production Department", "🏭")
    elif selected == "Maintenance":
        render_maintenance_subpages()
    elif selected == "Utility":
        render_utility_subpages()
    elif selected == "Power Generation":
        render_kpi_page("POWER GENERATION", "Power Generation", "⚡")
    elif selected == "Quality Control":
        render_kpi_page("QUALITY CONTROL", "Quality Control", "🔬")
    elif selected == "Quarry & Crusher":
        render_kpi_page("QUARRY & CRUSHER", "Quarry & Crusher", "⛏️")
    elif selected == "VRM Advanced ML":
        render_vrm_advanced()
    elif selected == "Kiln ML":
        render_kiln_ml()

if __name__ == "__main__":
    main()